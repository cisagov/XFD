"""Scripts to prune auto-enumerated Flare assets."""

# Standard Python Libraries
import asyncio
import datetime
import logging
import os
import socket
import time
import traceback

# Third-Party Libraries
import aioping
import numpy as np
import openpyxl
from openpyxl import load_workbook
import pandas as pd
import requests

# cisagov Libraries
from pe_source.data.config_source import create_retry_session
from pe_source.data.db_query_source import get_orgs
from pe_source.flare_ident_prune.flare_helpers import get_flare_token

# Set up logging
LOGGER = logging.getLogger(__name__)


def flare_identifiers_endpoint(token, params):
    """Call the Flare get identifiers endpoint with the specified parameters."""
    # Setup API call
    session = create_retry_session()
    url = "https://api.flare.io/firework/v3/identifiers/"
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    # Make API Call
    try:
        response = session.get(url, headers=headers, params=params, timeout=60)
        response.raise_for_status()
        return token, response
    except requests.exceptions.HTTPError as http_err:
        # Catch special token expiry scenario
        if response.status_code == 401:
            return token, response
        LOGGER.error(f"HTTP error occurred: {http_err}")
    except requests.exceptions.ConnectionError as conn_err:
        LOGGER.error(f"Connection error occurred: {conn_err}")
    except requests.exceptions.Timeout as timeout_err:
        LOGGER.error(f"Timeout error occurred: {timeout_err}")
    except requests.exceptions.RequestException as err:
        LOGGER.error(f"Unexpected error occurred: {err}")
    return token, None


def parse_domain_idents(raw_resp):
    """Parse out domain identifiers given raw API response."""
    resp = raw_resp.json()
    next = resp.get("next")
    total_ct = resp.get("total_count")
    # Get domains
    domain_list = []
    resp_list = resp.get("items")
    for ident in resp_list:
        domain_dict = {
            "id": ident.get("id"),
            "type": ident.get("type"),
            "value": ident.get("name"),
            "ip": None,
            "source": ident.get("source"),
            "curr_enabled": not ident.get("is_disabled"),
            "detected_resolvable": False,
        }
        domain_list.append(domain_dict)
    # Return parsed results
    return {
        "domains": domain_list,
        "next": next,
        "total_count": total_ct,
    }


def get_all_autoenum_domains(custom_params={}):
    """Get Flare auto-enumerated subdomains across all organizations."""
    all_domain_list = []
    chunk_size = 100  # Max is 100
    flare_token = get_flare_token()
    next = None
    # Make initial API call
    base_params = {
        "source_group": "SYSTEM",
        "types": ["domain"],
        "size": chunk_size,
    }
    params = base_params | custom_params
    flare_token, ini_resp = flare_identifiers_endpoint(flare_token, params)
    # Parse domain identifiers
    ini_resp_dict = parse_domain_idents(ini_resp)
    next = ini_resp_dict.get("next")
    all_domain_list.extend(ini_resp_dict.get("domains"))
    total_ident_count = ini_resp_dict.get("total_count")
    print(
        f"Retrieved {len(all_domain_list)} of {total_ident_count} auto-enum identifiers"
    )
    # If there's a next value, continue retrieval
    while next is not None:
        # Make API call for this chunk
        curr_base_params = {
            "from": next,
            "source_group": "SYSTEM",
            "types": ["domain"],
            "size": chunk_size,
        }
        curr_params = curr_base_params | custom_params
        flare_token, curr_resp = flare_identifiers_endpoint(flare_token, curr_params)
        # 401 token refresh check
        if curr_resp.status_code == 401:
            LOGGER.warning("401 code encountered, refreshing token")
            flare_token = get_flare_token()
            flare_token, curr_resp = flare_identifiers_endpoint(
                flare_token, curr_params
            )
        # Parse domain identifiers
        curr_resp_dict = parse_domain_idents(curr_resp)
        next = curr_resp_dict.get("next")
        all_domain_list.extend(curr_resp_dict.get("domains"))
        print(
            f"Retrieved {len(all_domain_list)} of {total_ident_count} auto-enum identifiers"
        )
    # Return results
    print("All auto-enum identifiers retrieved")
    return all_domain_list


async def check_ip_reachable(ip, count=1, timeout=3.0):
    """Check if a single IP is reachable using specified ping count and timeout."""
    # Attempt to ping IP the specified amount of times
    for i in range(count):
        try:
            # Attempt ping
            delay = await aioping.ping(ip, timeout=timeout)
            # If response receieved, return results
            return {
                "ip": ip,
                "detected_reachable": True,
                "response_delay": delay,
            }
        except (TimeoutError, PermissionError) as e:
            # If ping failed, try again
            print(f"\tPing {i+1}/{count} failed for {ip} - {e}")
    # If all ping attempts fail, mark as unreachable
    return {
        "ip": ip,
        "detected_reachable": False,
        "response_delay": None,
    }


async def check_ip_list_reachable(ip_list):
    """Launch multiple tasks to check IPs' reachability."""
    # Create separate tasks for each IP
    ping_ct = 5
    timeout = 3
    tasks = [check_ip_reachable(ip, ping_ct, timeout) for ip in ip_list]
    results = await asyncio.gather(*tasks)
    return results


def check_domains_responsive(domain_list):
    """Check each domain in list to see if it's resolvable/reachable."""
    domain_df = pd.DataFrame(domain_list)
    # Check resolvability of each domain
    for (
        idx,
        row,
    ) in domain_df.iterrows():
        domain = row["value"]
        # Test if domain has an IP associated with it (resolvable)
        print(
            f'Checking resolvability of domain "{domain}" ({idx+1} of {len(domain_df)})'
        )
        try:
            domain_ip = socket.gethostbyname(domain)
            resolvable = True
        except socket.gaierror:
            domain_ip = None
            resolvable = False
        # Update value in this row
        domain_df.at[idx, "ip"] = domain_ip
        domain_df.at[idx, "detected_resolvable"] = resolvable
    # Check reachability of any domains that have an IP (resolvable)
    ip_list = list(set(domain_df.loc[domain_df["ip"].notnull()]["ip"]))
    ip_results_df = pd.DataFrame(asyncio.run(check_ip_list_reachable(ip_list)))
    # Join resolvability and reachability results
    domain_df = pd.merge(domain_df, ip_results_df, on="ip", how="left")
    domain_df["detected_reachable"] = domain_df["detected_reachable"].fillna(False)
    domain_df["response_delay"] = domain_df["response_delay"].fillna(-1)
    # Calculate overall responsiveness and required action
    domain_df["detected_responsive"] = (
        domain_df["detected_resolvable"] & domain_df["detected_reachable"]
    )
    conditions = [
        (domain_df["curr_enabled"]) & (~domain_df["detected_responsive"]),
        (~domain_df["curr_enabled"]) & (domain_df["detected_responsive"]),
    ]
    choices = ["DISABLE", "ENABLE"]
    domain_df["required_action"] = np.select(conditions, choices, default="NO ACTION")
    domain_df = domain_df[
        [
            "id",
            "type",
            "value",
            "ip",
            "source",
            "curr_enabled",
            "detected_resolvable",
            "detected_reachable",
            "detected_responsive",
            "required_action",
        ]
    ]
    # Return results
    enable_list = domain_df.loc[domain_df["required_action"] == "ENABLE"].to_dict(
        orient="records"
    )
    disable_list = domain_df.loc[domain_df["required_action"] == "DISABLE"].to_dict(
        orient="records"
    )
    return enable_list, disable_list, domain_df


def toggle_ident(token, ident_id, active=True):
    """Enable or disable the Flare identifier based on specified ID."""
    # Setup API call
    session = create_retry_session()
    url = f"https://api.flare.io/firework/v2/assets/{ident_id}/toggle"
    headers = {"Authorization": f"Bearer {token}", "Content-Type": "application/json"}
    payload = {"is_disabled": not active}
    # Make API Call
    try:
        response = session.post(url, json=payload, headers=headers, timeout=60)
        response.raise_for_status()
        return token, response
    except requests.exceptions.HTTPError as http_err:
        # Catch special token expiry scenario
        if response.status_code == 401:
            return token, response
        LOGGER.error(f"HTTP error occurred: {http_err}")
    except requests.exceptions.ConnectionError as conn_err:
        LOGGER.error(f"Connection error occurred: {conn_err}")
    except requests.exceptions.Timeout as timeout_err:
        LOGGER.error(f"Timeout error occurred: {timeout_err}")
    except requests.exceptions.RequestException as err:
        LOGGER.error(f"Unexpected error occurred: {err}")
    return token, None


def update_ident_lists(enable_list, disable_list):
    """Enable/Disable the provided lists of Flare identifiers."""
    # Iterate over each identifier in enable list
    if len(enable_list) > 0:
        token = get_flare_token()
        for idx, ident in enumerate(enable_list):
            # Call endpoint to enable indentifier
            curr_ident_id = ident.get("id")
            curr_ident_name = ident.get("value")
            token, resp = toggle_ident(token, curr_ident_id, active=True)
            # General error handling
            if resp is None:
                LOGGER.error(
                    f'Failed to enable identifier "{curr_ident_name}" ({curr_ident_id}) {idx+1} of {len(disable_list)}, skipping...'
                )
                continue
            # 401 token refresh check
            if resp.status_code == 401:
                LOGGER.warning("401 code encountered, refreshing token")
                token = get_flare_token()
                token, resp = toggle_ident(token, curr_ident_id, active=True)
            print(
                f'Enabled identifier "{curr_ident_name}" ({curr_ident_id}) {idx+1} of {len(enable_list)}'
            )
        LOGGER.info("All identifiers marked for re-enabling have been re-enabled")
    else:
        LOGGER.info("No disabled identifiers to re-enable, continuing")
    # Iterate over each identifier in disable list
    if len(disable_list) > 0:
        token = get_flare_token()
        for idx, ident in enumerate(disable_list):
            curr_ident_id = ident.get("id")
            curr_ident_name = ident.get("value")
            # Call endpoint to disable indentifier
            token, resp = toggle_ident(token, curr_ident_id, active=False)
            # General error handling
            if resp is None:
                LOGGER.error(
                    f'Failed to disable identifier "{curr_ident_name}" ({curr_ident_id}) {idx+1} of {len(disable_list)}, skipping...'
                )
                continue
            # 401 token refresh check
            if resp.status_code == 401:
                LOGGER.warning("401 code encountered, refreshing token")
                token = get_flare_token()
                token, resp = toggle_ident(token, curr_ident_id, active=False)
            print(
                f'Disabled identifier "{curr_ident_name}" ({curr_ident_id}) {idx+1} of {len(disable_list)}'
            )
        LOGGER.info("All identifiers marked for disabling have been disabled")
    else:
        LOGGER.info("No enabled identifiers to disable, continuing")


def run_flare_ident_prune(orgs_list):
    """Prune flare auto-enumerated assets."""
    # Retrieve full org info from PE database
    pe_orgs = get_orgs()
    pe_orgs_final = []
    if orgs_list == "all":
        for pe_org in pe_orgs:
            if pe_org["report_on"]:
                pe_orgs_final.append(pe_org)
            else:
                continue
    elif orgs_list == "DEMO":
        for pe_org in pe_orgs:
            if pe_org["demo"]:
                pe_orgs_final.append(pe_org)
            else:
                continue
    else:
        for org in orgs_list:
            org_dict = next((d for d in pe_orgs if d["cyhy_db_name"] == org), None)
            pe_orgs_final.append(org_dict)
    # Alphabetize org list for consistent order
    pe_orgs_final = sorted(pe_orgs_final, key=lambda d: d["cyhy_db_name"])
    # Create file for exe time performance logging
    pe_orgs_final_df = pd.DataFrame(pe_orgs_final)
    current_date = datetime.date.today().strftime("%Y-%m-%d")
    first_org = pe_orgs_final_df.iloc[0]["cyhy_db_name"]
    last_org = pe_orgs_final_df.iloc[-1]["cyhy_db_name"]
    exe_time_file = (
        os.path.dirname(os.path.abspath(__file__))
        + f"/exe_time_logs/flare_prune_logs/flare_prune_{current_date}_{first_org}-{last_org}_exe_times.xlsx"
    )
    if not os.path.exists(exe_time_file):
        workbook = openpyxl.Workbook()
        sheet = workbook["Sheet"]
        sheet.append(
            [
                "timestamp",
                "org_abbrv",
                "exe_time",
            ]
        )
        workbook.save(exe_time_file)

    # Begin Flare asset prune
    time_start = time.time()
    try:
        # Retrieve list of all auto-enum assets in Flare
        LOGGER.info("Retrieving all auto-enumerated assets within Flare")
        curr_enabled_domains = get_all_autoenum_domains(
            custom_params={"is_disabled": False}
        )
        LOGGER.info("All auto-enumerated assets that are currently enabled retrieved")
        curr_disabled_domains = get_all_autoenum_domains(
            custom_params={"is_disabled": True}
        )
        LOGGER.info("All auto-enumerated assets that are currently disabled retrieved")
        auto_enum_domains = curr_enabled_domains + curr_disabled_domains
        LOGGER.info("All auto-enumerated assets retrieved")
        # Don't modify certain domains
        excluded_roots = (
            "doj.gov",
            "nrc-gateway.gov",
            "nrc.gov",
            "cisa.dhs.gov",
        )
        auto_enum_domains = [
            d for d in auto_enum_domains if not d["value"].endswith(excluded_roots)
        ]
        # Check which domains are responsive
        LOGGER.info("Checking which auto-enumerated assets are responsive")
        enable_list, disable_list, all_resp_results = check_domains_responsive(
            auto_enum_domains
        )
        LOGGER.info("Auto-enumerated assets have been checked for responsiveness")
        # Enable/Disable the appropriate domains
        LOGGER.info("Enabling/Disabling auto-enumerated assets based on responsiveness")
        update_ident_lists(enable_list, disable_list)
        LOGGER.info(
            "Auto-enumerated assets have been enabled/disabled based on responsiveness"
        )
        # Save responsiveness results to file
        prune_results_file = (
            os.path.dirname(os.path.abspath(__file__))
            + f"/flare_prune_results/flare_prune_results_{current_date}.xlsx"
        )
        all_resp_results.to_excel(prune_results_file, index=False)
        # Print stats
        total_assets_tested = len(all_resp_results)
        total_enable_assets = len(
            all_resp_results.loc[all_resp_results["required_action"] == "ENABLE"]
        )
        total_disable_assets = len(
            all_resp_results.loc[all_resp_results["required_action"] == "DISABLE"]
        )
        post_prune_total = (
            total_assets_tested + total_enable_assets - total_disable_assets
        )
        LOGGER.info(f"Total Flare Assets Tested: {total_assets_tested}")
        LOGGER.info(f"Total Flare Assets Enabled: {total_enable_assets}")
        LOGGER.info(f"Total Flare Assets Disabled: {total_disable_assets}")
        LOGGER.info(f"Total Flare Asssets Post Pruning: {post_prune_total}")
    except Exception as e:
        LOGGER.error(f"Encountered an error during Flare pruning script - {e}")
        traceback.print_exc()

    # Write exe time to file
    time_end = time.time()
    org_exe_time = "{:.5f}".format(
        datetime.timedelta(seconds=(time_end - time_start)).total_seconds()
    )
    org_exe_stats = [
        str(datetime.datetime.now()),
        "ALL ORGS",
        org_exe_time,
    ]
    workbook = load_workbook(exe_time_file)
    sheet = workbook["Sheet"]
    sheet.append(org_exe_stats)
    workbook.save(exe_time_file)


# --- Additional functions for troubleshooting ---
def get_sys_assets_by_enable_status(enabled=True):
    """Get Flare system-added assets across all organizations based on enable/disable status."""
    is_disabled = not enabled
    all_ident_list = []
    chunk_size = 100
    flare_token = get_flare_token()
    next = None
    # Make initial API call
    params = {
        "source_group": "SYSTEM",
        "types": ["domain"],
        "size": chunk_size,
        "is_disabled": is_disabled,
    }
    flare_token, ini_resp = flare_identifiers_endpoint(flare_token, params)
    # Parse domain identifiers
    ini_resp_dict = parse_domain_idents(ini_resp)
    next = ini_resp_dict.get("next")
    total_ident_count = ini_resp_dict.get("total_count")
    all_ident_list.extend(ini_resp_dict.get("domains"))
    print(f"Retrieved {len(all_ident_list)} of {total_ident_count} system identifiers")
    # If there's a next value, continue retrieval
    while next is not None:
        # Make API call for this chunk
        curr_params = {
            "from": next,
            "source_group": "SYSTEM",
            "types": ["domain"],
            "size": chunk_size,
            "is_disabled": is_disabled,
        }
        flare_token, curr_resp = flare_identifiers_endpoint(flare_token, curr_params)
        # 401 token refresh check
        if curr_resp.status_code == 401:
            LOGGER.warning("401 code encountered, refreshing token")
            flare_token = get_flare_token()
            flare_token, curr_resp = flare_identifiers_endpoint(
                flare_token, curr_params
            )
        # Parse domain identifiers
        curr_resp_dict = parse_domain_idents(curr_resp)
        next = curr_resp_dict.get("next")
        all_ident_list.extend(curr_resp_dict.get("domains"))
        print(
            f"Retrieved {len(all_ident_list)} of {total_ident_count} system identifiers"
        )
    # Return results
    print("All system identifiers retrieved")
    return all_ident_list


def toggle_ident_list(ident_list, action):
    """Enable/Disable the provided lists of Flare identifiers."""
    # Iterate over each identifier in enable list
    if action == "ENABLE":
        token = get_flare_token()
        for idx, ident in enumerate(ident_list):
            # Call endpoint to enable indentifier
            curr_ident_id = ident.get("id")
            curr_ident_name = ident.get("value")
            token, resp = toggle_ident(token, curr_ident_id, active=True)
            # 401 token refresh check
            if resp.status_code == 401:
                print("401 code encountered, refreshing token")
                token = get_flare_token()
                token, resp = toggle_ident(token, curr_ident_id, active=True)
            print(
                f'Enabled identifier "{curr_ident_name}" ({curr_ident_id}) {idx+1} of {len(ident_list)}'
            )
        print("All identifiers in list have been enabled")
    elif action == "DISABLE":
        token = get_flare_token()
        for idx, ident in enumerate(ident_list):
            # Call endpoint to disable indentifier
            curr_ident_id = ident.get("id")
            curr_ident_name = ident.get("value")
            token, resp = toggle_ident(token, curr_ident_id, active=False)
            # 401 token refresh check
            if resp.status_code == 401:
                print("401 code encountered, refreshing token")
                token = get_flare_token()
                token, resp = toggle_ident(token, curr_ident_id, active=False)
            print(
                f'Disabled identifier "{curr_ident_name}" ({curr_ident_id}) {idx+1} of {len(ident_list)}'
            )
        print("All identifiers in list have been disabled")


def toggle_all_sys_assets(action):
    """Toggle all system-added Flare assets on or off."""
    os.environ["FLARE_KEY_NUM"] = "1"
    if action == "ENABLE_ALL":
        print(">>> Enabling all Flare system assets <<<")
        # Retrieve all currently disabled assets
        print("Retrieving list of all currently disabled system assets...")
        disabled_asset_list = get_sys_assets_by_enable_status(False)
        # Enable those disabled assets
        print("Enabling all assets that are currently disabled....")
        toggle_ident_list(disabled_asset_list, "ENABLE")
        print("The following identifiers were enabled:")
        print(pd.DataFrame(disabled_asset_list))
    elif action == "DISABLE_ALL":
        print(">>> Disabling all Flare system assets <<<")
        # Retrueve all curently enabled assets
        print("Retrieving list of all currently enabled system assets...")
        enabled_asset_list = get_sys_assets_by_enable_status(True)
        # Disable those enabled assets
        print("Disabling all assets that are currently enabled....")
        # toggle_ident_list(enabled_asset_list, "DISABLE")
        print("The following identifiers were disabled:")
        print(pd.DataFrame(enabled_asset_list))
    else:
        print('ERROR: Action not recognized, options are "ENABLE_ALL" or "DISABLE_ALL"')
