#!/usr/bin/env python3
from configparser import ConfigParser
from openpyxl import load_workbook
import docker
import os
import subprocess
import time
from datetime import date, timedelta
import time
from pathlib import Path
import shutil
from PyPDF4 import PdfFileReader, utils
from was_dynamodb.dynamodb import DynamoDB
from os import environ
from typing import Any
import zipfile
from datetime import datetime

CUSTOMER_DATA = DynamoDB(
    environ.get("STAKEHOLDERS_TABLE_NAME"), environ.get("PROD_PROFILE")
)

today = date.today()
today = today - timedelta(hours=4)
dateToday = today.strftime("%Y%m%d")
pdf_date = today.strftime('%Y-%m-%d')

cwd = os.getcwd()

config = ConfigParser()
config.read(Path(__file__).resolve().parent / 'was_config.txt')
dailyReportsFilePath = config.get('was_files', 'dailyReportsFilePath')
powerAutomateTriggers = config.get('was_files', 'powerAutomateTriggers')

def main():
    empty_folder(powerAutomateTriggers)
    # Load excel workbooks and set active sheets
    dailyReports = load_workbook(filename=dailyReportsFilePath)
    dailyReportsSheet = dailyReports.active
    report_time_start = time.perf_counter()
    empty_cell_count = 0
    MAX_CONCURRENT_CONTAINERS = 30
    global current_containers
    current_containers = 0
    completed_containers = 0
    
    for cell in dailyReportsSheet['B']:
        password = cell.offset(row=0,column=16).value
        if cell.offset(row=0,column=3).value == 'Running':
            continue
        if cell.offset(row=0,column=5).value is not None:
            continue
        if cell.value is None:
            empty_cell_count += 1
            if empty_cell_count == 2:
                break
            else:
                continue
        else:
            while current_containers >= MAX_CONCURRENT_CONTAINERS:
                time.sleep(5)
                previous_container_count = current_containers
                current_containers = get_running_container_count()
                completed_containers += (previous_container_count - current_containers)
                print(completed_containers, 'done so far.')
                
            # Start the container
            run_container(cell.value, dateToday, cwd, password)
        
    while get_running_container_count() != 0:
        time.sleep(5)
        previous_container_count = current_containers
        current_containers = get_running_container_count()
        completed_containers += (previous_container_count - current_containers)
        print(completed_containers, 'done so far.')
    
    print('Reports are all done!', completed_containers, 'generated.')
    report_time_end = time.perf_counter()
    print(f"Total Report Generation Time: {report_time_end - report_time_start:0.0f} seconds")
    
    moveReports(copyReports())
    zip_folder(powerAutomateTriggers)

def run_container(tag, dateToday, cwd, password):
    if password == 'STATIC PASSWORD':
        password = get_static_password(tag)
    if len(password) > 1 and "'" in password[1:-1]:
        return
    global current_containers
    client = docker.from_env()
    try:
        client.containers.run('was_report_creator', 'python3 WAS_report_creator.py -t ' + tag + ' --encrypt ' + password, remove=False, stdin_open=True, tty=True, name=tag + '_' + dateToday, volumes=[str(Path.home()) + ':/WAS_REPORT_GENERATION/docs'], detach=True) ### restart_policy={"Name": "on-failure", "MaximumRetryCount": 5}
    except:
        client.containers.run('was_report_creator', 'python3 WAS_report_creator.py -t ' + tag + ' --encrypt ' + password, remove=False, stdin_open=True, tty=True, volumes=[str(Path.home()) + ':/WAS_REPORT_GENERATION/docs'], detach=True) ### restart_policy={"Name": "on-failure", "MaximumRetryCount": 5}
    # try:
    #     client.containers.run('was_report_creator', 'python3 WAS_report_creator.py -t ' + tag + ' --encrypt ' + password, 
    #                       remove=False, stdin_open=True, tty=True, 
    #                       name=tag + '_' + dateToday, 
    #                       volumes=[str(cwd) + ':/WAS_REPORT_GENERATION/docs'], detach=True) 
    # ### restart_policy={"Name": "on-failure", "MaximumRetryCount": 5}
    # except:
    #     client.containers.run('was_report_creator', 'python3 WAS_report_creator.py -t ' + tag + ' --encrypt ' + password, 
    #                       remove=False, stdin_open=True, tty=True, 
    #                       volumes=[str(cwd) + ':/WAS_REPORT_GENERATION/docs'], detach=True) 
    ### restart_policy={"Name": "on-failure", "MaximumRetryCount": 5}
    current_containers += 1
    
def get_running_container_count():
    s = subprocess.check_output('docker ps', shell=True)
    return s.count(b'\n') - 1  # Subtract one for the header row

def copyReports():
    print('Copying reports to Teams Folder...')
    homeDir = Path.home()

    dailyReports = load_workbook(filename=dailyReportsFilePath)
    dailyReportsSheet = dailyReports.active

    row = 1
    filenames = []
    empty_cell_count = 0

    for cell in dailyReportsSheet['B']:
        if cell.offset(row=0,column=3).value == 'Running':
            continue
        if cell.offset(row=0,column=5).value is not None:
            continue
        if cell.value is None:
            empty_cell_count += 1
            if empty_cell_count == 2:
                break
            else:
                continue
        filename = cell.value + '_report_' + pdf_date + '.pdf'
        src = str(homeDir) + '/' + filename
        if is_damaged(src) or "UNENCRYPTED" in src:
            # print(f"DAMAGED FILE, REGENERATE: {filename}")
            continue
        filenames.append(filename)
        
        # dst = dailyReportsFilePath.rsplit('/', 1)[0] + '/Report Mailer'
        dst = powerAutomateTriggers
        try:
            shutil.copy(src, dst)
        except:
            print('Error copying ' + filename)
            row += 1
            continue
        row += 1
    
    # file = open(str(powerAutomateTriggers) + '/send_reports.txt', 'w')
    # file.write('send_reports')
    # file.close()

    return filenames

def moveReports(filenames):
    print('Moving reports to reports archive...')
    homeDir = Path.home()

    for file in filenames:
        src = str(homeDir) + '/' + file
        dst = str(homeDir) + '/Reports'
        try:
            shutil.move(src, dst)
        except:
            print('Error moving ' + file)
            continue

def is_damaged(filepath):
    try:
        if not os.path.exists(filepath):
            raise FileNotFoundError(f"File not found: {filepath}")
        
        with open(filepath, 'rb') as pdf:
            reader = PdfFileReader(pdf)
            if reader.isEncrypted:
                return False
            else:
                print("report unencrypted. will not copy")
                return True
    except FileNotFoundError as e:
        print("FILE NOT FOUND")
        return False
    except Exception as e:
        print(f"Error: {e}")
        print(f"DAMAGED FILE, REGENERATE: {filepath}")
        return True

def get_static_password(tag):
    try:
        stakeholder = CUSTOMER_DATA.get_item({"Tag": {"S": tag}})
    except KeyError:
        tag_split = tag.split('_')[0]
        stakeholder = CUSTOMER_DATA.get_item({"Tag": {"S": tag_split}})
    static_password = get_dynamo_value(stakeholder, "Report Password")
    return f"'{static_password}'"


def get_dynamo_value(stakeholder: dict, key: str) -> Any:
    """
    gets specific dynamo values

    Parameters
    ----------
    stakeholder : dict
        stakeholder dynamo db dict
    key : str
        stakeholder attribute

    Returns
    -------
    value
        value of field
    """
    if key in stakeholder:
        return list(stakeholder[key].values())[0]
    return None


def zip_folder(folder_path):
    """
    Zips all files in the given folder into a single zip archive (saved inside the same folder),
    then deletes the original files after successful compression.
    The zip file name includes the 4-digit date (e.g. reports1024.zip for October 24).
    """
    # Get current date as MMDD (e.g., 1024 for October 24)
    date_str = datetime.now().strftime("%m%d")

    # Get folder name (e.g., 'reports' from 'path/to/reports')
    # folder_name = os.path.basename(os.path.normpath(folder_path))

    # Path for the zip file (inside the same folder)
    zip_filename = f"reports{date_str}.zip"
    zip_path = os.path.join(folder_path, zip_filename)

    # Create the zip file
    with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
        for filename in os.listdir(folder_path):
            file_path = os.path.join(folder_path, filename)
            # Include only files (skip the zip itself if rerun)
            if os.path.isfile(file_path) and filename != zip_filename:
                zipf.write(file_path, arcname=filename)

    # Delete the original files after successful zipping
    for filename in os.listdir(folder_path):
        file_path = os.path.join(folder_path, filename)
        if os.path.isfile(file_path) and filename != zip_filename:
            os.remove(file_path)

    print(f"Created zip file: {zip_path}")
    print(f"Deleted original files inside: {folder_path}")
    return zip_path

import os
import shutil
import zipfile
from datetime import datetime

def empty_folder(folder_path):
    """
    Deletes all files and subfolders inside the specified folder,
    but keeps the folder itself.
    """
    for item in os.listdir(folder_path):
        item_path = os.path.join(folder_path, item)
        if os.path.isfile(item_path) or os.path.islink(item_path):
            os.remove(item_path)
        elif os.path.isdir(item_path):
            shutil.rmtree(item_path)
    print(f"Emptied folder: {folder_path}")


if __name__ == '__main__':
    main()
